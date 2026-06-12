#!/usr/bin/env python3
"""Rename sonrası Excel güncellemesi: Kod Eşleme'ye satır ekle, aksiyon durumlarını işle."""
import json
import os
import unicodedata
from openpyxl import load_workbook

HOME = os.environ['HOME']
EXCEL = os.path.join(HOME, 'Desktop/KademeQMS_Dokumanlar/00_KYS_Analiz/Revizyon/Kademe_KYS_Aksiyon_Takip Revizyon.xlsx')
WORKDIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(WORKDIR, 'plan.json'), encoding='utf-8') as f:
    data = json.load(f)
plan = data['plan']
new_mappings = data['newMappings']

wb = load_workbook(EXCEL)

# ── 1) Kod Eşleme sayfasına ekle ─────────────────────────────────────────────
ws = wb['Kod Eşleme']
# başlık satırını bul (Eski Kod)
header_row = None
for r in range(1, 10):
    if str(ws.cell(row=r, column=1).value or '').strip() == 'Eski Kod':
        header_row = r
        break
if header_row is None:
    raise SystemExit('Kod Eşleme başlık satırı bulunamadı')

last = ws.max_row
existing = set()
for r in range(header_row + 1, last + 1):
    v = ws.cell(row=r, column=1).value
    if v:
        existing.add(unicodedata.normalize('NFC', str(v).strip()))

added = 0
row = last + 1
for m in new_mappings:
    old = unicodedata.normalize('NFC', str(m['old']))
    if old in existing:
        continue
    ws.cell(row=row, column=1, value=m['old'])
    ws.cell(row=row, column=2, value=m['neu'])
    ws.cell(row=row, column=3, value=m['folder'])
    ws.cell(row=row, column=4, value=m['oldFile'])
    ws.cell(row=row, column=5, value=m['newFile'])
    ws.cell(row=row, column=6, value=f"Otomatik rename 12.06.2026 — {m['note']}")
    row += 1
    added += 1

# ── 2) Aksiyon Listesi durumları ─────────────────────────────────────────────
wsa = wb['Aksiyon Listesi']
# kolonlar: No=1 ... Durum, Kapanış Notu son iki kolon — başlıktan bul
hdr = {}
for c in range(1, wsa.max_column + 1):
    hdr[str(wsa.cell(row=1, column=c).value or '').strip()] = c
col_no = hdr['No']
col_durum = hdr['Durum']
col_not = hdr['Kapanış Notu']

# plan'daki aksiyon numaraları
acted = set()
for p in plan:
    for no in p['no'].split('+'):
        acted.add(no)

NOTE = "Dosya adı 12.06.2026'da otomatik düzeltildi (rename_log_2026-06-12.csv); antet ve kaynak dosya içeriği güncellenmeli."
NOTE_DONE = "Dosya adları 12.06.2026'da otomatik düzeltildi (rename_log_2026-06-12.csv)."

updated = 0
for r in range(2, wsa.max_row + 1):
    no = str(wsa.cell(row=r, column=col_no).value or '').strip()
    if no not in acted:
        continue
    cur = str(wsa.cell(row=r, column=col_durum).value or '').strip()
    if cur == 'Tamamlandı':
        continue
    if no == 'A200':
        wsa.cell(row=r, column=col_durum, value='Tamamlandı')
        wsa.cell(row=r, column=col_not, value=NOTE_DONE)
    else:
        wsa.cell(row=r, column=col_durum, value='Devam Ediyor')
        wsa.cell(row=r, column=col_not, value=NOTE)
    updated += 1

wb.save(EXCEL)
print(f"Kod Eşleme: +{added} satır | Aksiyon durumu güncellenen: {updated}")
